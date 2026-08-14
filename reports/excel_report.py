import pandas as pd
import os
from datetime import datetime

def create_capital_allocation_sheet(
    writer,
    capital_allocation
):
    """
    Create the Capital Allocation worksheet.

    Expected structure from generate_capital_allocation():

        {
            "Capital Summary": DataFrame,
            "Capital Allocation": DataFrame
        }

    The allocation DataFrame contains:
        BUY NEW
        BUY MORE
        REDUCE %
        SELL
        HOLD
    """

    print("Creating Capital Allocation")

    if capital_allocation is None:
        print("No capital allocation data available")
        return

    if not isinstance(capital_allocation, dict):
        print(
            "Capital allocation has unexpected type:",
            type(capital_allocation)
        )
        return

    sheet = "Capital Allocation"
    row = 0

    # =========================================================
    # GET DATA FROM CURRENT CAPITAL ALLOCATOR STRUCTURE
    # =========================================================

    summary_df = capital_allocation.get(
        "Capital Summary",
        pd.DataFrame()
    )

    allocation_df = capital_allocation.get(
        "Capital Allocation",
        pd.DataFrame()
    )

    if summary_df is None:
        summary_df = pd.DataFrame()

    if allocation_df is None:
        allocation_df = pd.DataFrame()

    if not isinstance(summary_df, pd.DataFrame):
        summary_df = pd.DataFrame(summary_df)

    if not isinstance(allocation_df, pd.DataFrame):
        allocation_df = pd.DataFrame(allocation_df)

    # =========================================================
    # HELPER
    # =========================================================

    def write_section(
        title,
        dataframe,
        current_row
    ):

        pd.DataFrame(
            {
                "Section": [title]
            }
        ).to_excel(
            writer,
            sheet_name=sheet,
            startrow=current_row,
            index=False
        )

        current_row += 2

        if (
            dataframe is not None
            and not dataframe.empty
        ):

            dataframe.to_excel(
                writer,
                sheet_name=sheet,
                startrow=current_row,
                index=False
            )

            current_row += len(dataframe) + 3

        else:

            pd.DataFrame(
                {
                    "Information": ["None"]
                }
            ).to_excel(
                writer,
                sheet_name=sheet,
                startrow=current_row,
                index=False
            )

            current_row += 4

        return current_row

    # =========================================================
    # CAPITAL SUMMARY
    # =========================================================

    row = write_section(
        "CAPITAL SUMMARY",
        summary_df,
        row
    )

    # =========================================================
    # SAFELY FILTER ALLOCATIONS
    # =========================================================

    if (
        not allocation_df.empty
        and
        "Action" in allocation_df.columns
    ):

        buy_new_df = allocation_df[
            allocation_df["Action"] == "BUY NEW"
        ].copy()

        buy_more_df = allocation_df[
            allocation_df["Action"] == "BUY MORE"
        ].copy()

        reduce_df = allocation_df[
            allocation_df["Action"].astype(str).str.startswith(
                "REDUCE"
            )
            |
            allocation_df["Action"].isin(
                ["SELL"]
            )
        ].copy()

        hold_df = allocation_df[
            allocation_df["Action"] == "HOLD"
        ].copy()

    else:

        buy_new_df = pd.DataFrame()
        buy_more_df = pd.DataFrame()
        reduce_df = pd.DataFrame()
        hold_df = pd.DataFrame()

    # =========================================================
    # BUY NEW
    # =========================================================

    row = write_section(
        "BUY NEW RECOMMENDATIONS",
        buy_new_df,
        row
    )

    # =========================================================
    # BUY MORE
    # =========================================================

    row = write_section(
        "BUY MORE RECOMMENDATIONS",
        buy_more_df,
        row
    )

    # =========================================================
    # REDUCE / SELL
    # =========================================================

    row = write_section(
        "REDUCE / SELL RECOMMENDATIONS",
        reduce_df,
        row
    )

    # =========================================================
    # HOLD
    # =========================================================

    row = write_section(
        "HOLD POSITIONS",
        hold_df,
        row
    )

    # =========================================================
    # ALL ACTIONS
    # =========================================================

    row = write_section(
        "ALL CAPITAL ALLOCATION ACTIONS",
        allocation_df,
        row
    )

    # =========================================================
    # BASIC FORMATTING
    # =========================================================

    try:

        workbook = writer.book
        worksheet = writer.sheets[sheet]

        # Works with openpyxl
        if hasattr(worksheet, "column_dimensions"):

            widths = {
                "A": 18,
                "B": 16,
                "C": 18,
                "D": 14,
                "E": 14,
                "F": 14,
                "G": 14,
                "H": 18,
                "I": 16,
                "J": 18,
                "K": 14,
                "L": 14,
                "M": 14,
                "N": 20,
                "O": 42,
                "P": 16,
                "Q": 16,
                "R": 16,
            }

            for column, width in widths.items():
                worksheet.column_dimensions[column].width = width

    except Exception as e:

        print(
            f"Capital Allocation formatting warning: {e}"
        )

    print("Capital Allocation sheet created")

def create_report(
    results,
    portfolio_summary,
    alerts,
    sector_summary,
    portfolio_actions,
    portfolio_optimisation,
    rebalance_recommendations,
    portfolio_health,
    capital_allocation,
    decisions,
    trade_plan,
    performance_summary=None,
    signal_performance=None,
    horizon_performance=None,
    score_performance=None,
    score_bucket_performance=None,
    component_score_performance=None,
    signal_horizon_performance=None,
    recommendation_intelligence=None,
    portfolio_ai_review=None,
    portfolio_manager_review=None,
    growth_plan=None,
    final_portfolio_decisions=None,
    recommendation_learning=None
):

    print("STARTING REPORT CREATION")

    # =========================================================
    # SAFETY HANDLING
    # =========================================================

    if results is None:
        results = []

    # Convert None values to safe DataFrames where appropriate
    if portfolio_summary is None:
        portfolio_summary = pd.DataFrame()

    if alerts is None:
        alerts = pd.DataFrame()

    if sector_summary is None:
        sector_summary = pd.DataFrame()

    if portfolio_actions is None:
        portfolio_actions = pd.DataFrame()

    if portfolio_optimisation is None:
        portfolio_optimisation = pd.DataFrame()

    if rebalance_recommendations is None:
        rebalance_recommendations = pd.DataFrame()

    if decisions is None:
        decisions = pd.DataFrame()

    if trade_plan is None:
        trade_plan = pd.DataFrame()

    if performance_summary is None:
        performance_summary = pd.DataFrame()

    if signal_performance is None:
        signal_performance = pd.DataFrame()

    if horizon_performance is None:
        horizon_performance = pd.DataFrame()

    if score_performance is None:
        score_performance = pd.DataFrame()

    if score_bucket_performance is None:
        score_bucket_performance = pd.DataFrame()

    if component_score_performance is None:
        component_score_performance = pd.DataFrame()

    if signal_horizon_performance is None:
        signal_horizon_performance = pd.DataFrame()

    if recommendation_intelligence is None:
        recommendation_intelligence = pd.DataFrame()

    if portfolio_ai_review is None:
        portfolio_ai_review = pd.DataFrame()

    if portfolio_manager_review is None:
        portfolio_manager_review = {}

    if growth_plan is None:
        growth_plan = []

    if final_portfolio_decisions is None:
        final_portfolio_decisions = pd.DataFrame()

    if recommendation_learning is None:
        recommendation_learning = {}

    # =========================================================
    # NORMALISE DATAFRAMES
    # =========================================================

    dataframe_names = [
        "portfolio_summary",
        "alerts",
        "sector_summary",
        "portfolio_actions",
        "portfolio_optimisation",
        "rebalance_recommendations",
        "decisions",
        "trade_plan",
        "performance_summary",
        "signal_performance",
        "horizon_performance",
        "score_performance",
        "score_bucket_performance",
        "component_score_performance",
        "signal_horizon_performance",
        "recommendation_intelligence",
        "portfolio_ai_review",
        "final_portfolio_decisions"
    ]

    for name in dataframe_names:

        value = locals().get(name)

        if not isinstance(value, pd.DataFrame):

            try:
                value = pd.DataFrame(value)
            except Exception:
                value = pd.DataFrame()

            locals()[name] = value

    # Explicit assignments because modifying locals() is unreliable
    if not isinstance(portfolio_summary, pd.DataFrame):
        portfolio_summary = pd.DataFrame(portfolio_summary)

    if not isinstance(alerts, pd.DataFrame):
        alerts = pd.DataFrame(alerts)

    if not isinstance(sector_summary, pd.DataFrame):
        sector_summary = pd.DataFrame(sector_summary)

    if not isinstance(portfolio_actions, pd.DataFrame):
        portfolio_actions = pd.DataFrame(portfolio_actions)

    if not isinstance(portfolio_optimisation, pd.DataFrame):
        portfolio_optimisation = pd.DataFrame(portfolio_optimisation)

    if not isinstance(rebalance_recommendations, pd.DataFrame):
        rebalance_recommendations = pd.DataFrame(
            rebalance_recommendations
        )

    if not isinstance(decisions, pd.DataFrame):
        decisions = pd.DataFrame(decisions)

    if not isinstance(trade_plan, pd.DataFrame):
        trade_plan = pd.DataFrame(trade_plan)

    if not isinstance(performance_summary, pd.DataFrame):
        performance_summary = pd.DataFrame(performance_summary)

    if not isinstance(signal_performance, pd.DataFrame):
        signal_performance = pd.DataFrame(signal_performance)

    if not isinstance(horizon_performance, pd.DataFrame):
        horizon_performance = pd.DataFrame(horizon_performance)

    if not isinstance(score_performance, pd.DataFrame):
        score_performance = pd.DataFrame(score_performance)

    if not isinstance(score_bucket_performance, pd.DataFrame):
        score_bucket_performance = pd.DataFrame(
            score_bucket_performance
        )

    if not isinstance(component_score_performance, pd.DataFrame):
        component_score_performance = pd.DataFrame(
            component_score_performance
        )

    if not isinstance(signal_horizon_performance, pd.DataFrame):
        signal_horizon_performance = pd.DataFrame(
            signal_horizon_performance
        )

    if not isinstance(recommendation_intelligence, pd.DataFrame):
        recommendation_intelligence = pd.DataFrame(
            recommendation_intelligence
        )

    if isinstance(portfolio_ai_review, list):
        portfolio_ai_review = pd.DataFrame(
            portfolio_ai_review
        )

    if not isinstance(portfolio_ai_review, pd.DataFrame):
        portfolio_ai_review = pd.DataFrame(
            portfolio_ai_review
        )

    if not isinstance(final_portfolio_decisions, pd.DataFrame):
        final_portfolio_decisions = pd.DataFrame(
            final_portfolio_decisions
        )

    # =========================================================
    # DEBUG INFORMATION
    # =========================================================

    print(
        "RECOMMENDATION INTELLIGENCE SHAPE:",
        recommendation_intelligence.shape
    )

    print(
        "ALERTS SHAPE:",
        alerts.shape
    )

    print(
        "RECOMMENDATION LEARNING TYPE:",
        type(recommendation_learning)
    )

    if isinstance(
        recommendation_learning,
        dict
    ):

        print(
            "RECOMMENDATION LEARNING KEYS:",
            list(
                recommendation_learning.keys()
            )
        )

    # =========================================================
    # FILENAME
    # =========================================================

    report_path = (
        "/Users/jameshulin/Documents/"
        "stock-momentum-agent/reports/"
    )

    os.makedirs(
        report_path,
        exist_ok=True
    )

    filename = os.path.join(
        report_path,
        f"daily_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )

    print("CREATING REPORT")

    print(
        "Stocks:",
        len(results)
    )

    print(
        "Portfolio:",
        len(portfolio_summary)
    )

    print(
        "Alerts:",
        len(alerts)
    )

    # =========================================================
    # CREATE WORKBOOK
    # =========================================================

    with pd.ExcelWriter(
        filename,
        engine="openpyxl"
    ) as writer:

        # =====================================================
        # EXECUTIVE SUMMARY
        # =====================================================

        print(
            "Creating Executive Summary"
        )

        summary_rows = [

            [
                "STOCK MOMENTUM AGENT - EXECUTIVE SUMMARY"
            ],

            [
                "Generated",
                datetime.now().strftime(
                    "%d %B %Y %H:%M"
                )
            ],

            [],

            [
                "PORTFOLIO OVERVIEW"
            ],

            [
                "Metric",
                "Value"
            ],

            [
                "Stocks Scanned",
                len(results)
            ],

            [
                "Portfolio Holdings",
                len(portfolio_summary)
            ],

            [
                "Alerts",
                len(alerts)
            ],

            [],

            [
                "PORTFOLIO HEALTH"
            ]

        ]

        # -----------------------------------------------------
        # Portfolio Health
        # -----------------------------------------------------

        if isinstance(
            portfolio_health,
            dict
        ):

            summary_rows.extend([

                [
                    "Health Score",
                    portfolio_health.get(
                        "Health Score",
                        ""
                    )
                ],

                [
                    "Rating",
                    portfolio_health.get(
                        "Rating",
                        ""
                    )
                ]

            ])

        # -----------------------------------------------------
        # Portfolio Manager Intelligence
        # -----------------------------------------------------

        if isinstance(
            portfolio_manager_review,
            dict
        ):

            summary_rows.extend([

                [],

                [
                    "AI PORTFOLIO MANAGER VIEW"
                ],

                [
                    "Market View",
                    portfolio_manager_review.get(
                        "Market View",
                        ""
                    )
                ],

                [
                    "Portfolio Status",
                    portfolio_manager_review.get(
                        "Portfolio Status",
                        ""
                    )
                ],

                [
                    "AI Summary",
                    portfolio_manager_review.get(
                        "AI Summary",
                        ""
                    )
                ]

            ])

            summary_rows.append([])

            summary_rows.append(
                [
                    "KEY STRENGTHS"
                ]
            )

            for item in portfolio_manager_review.get(
                "Key Strengths",
                []
            ):

                summary_rows.append(
                    [
                        item
                    ]
                )

            summary_rows.append([])

            summary_rows.append(
                [
                    "KEY RISKS"
                ]
            )

            for item in portfolio_manager_review.get(
                "Key Risks",
                []
            ):

                summary_rows.append(
                    [
                        item
                    ]
                )

            summary_rows.append([])

            summary_rows.append(
                [
                    "PRIORITY ACTIONS"
                ]
            )

            for item in portfolio_manager_review.get(
                "Priority Actions",
                []
            ):

                summary_rows.append(
                    [
                        item
                    ]
                )

        # -----------------------------------------------------
        # Top Opportunities
        # -----------------------------------------------------

        summary_rows.extend([

            [],

            [
                "TOP OPPORTUNITIES"
            ],

            [
                "Ticker",
                "Signal",
                "Investment Score"
            ]

        ])

        if len(results) > 0:

            results_df = pd.DataFrame(
                results
            )

            if "Investment Score" in results_df.columns:

                opportunities = (
                    results_df
                    .sort_values(
                        "Investment Score",
                        ascending=False
                    )
                    .head(5)
                )

            else:

                opportunities = results_df.head(5)

            for _, stock in opportunities.iterrows():

                summary_rows.append(
                    [

                        stock.get(
                            "Ticker",
                            ""
                        ),

                        stock.get(
                            "Signal",
                            ""
                        ),

                        stock.get(
                            "Investment Score",
                            ""
                        )

                    ]
                )

        pd.DataFrame(
            summary_rows
        ).to_excel(
            writer,
            sheet_name="Executive Summary",
            index=False,
            header=False
        )

        # =====================================================
        # HOW TO USE
        # =====================================================

        print(
            "Creating How To Use"
        )

        pd.DataFrame(
            {
                "Instructions": [

                    "Executive Summary provides portfolio overview",

                    "Stock Rankings shows investment opportunities",

                    "Portfolio Actions shows recommended changes",

                    "Capital Allocation shows suggested buys, reductions, sector avoidance and cash deployment",

                    "Recommendation Intelligence measures historical recommendation quality",

                    "Recommendation Learning measures how recommendation quality changes across horizons, signals, scores and confidence levels"

                ]
            }

        ).to_excel(
            writer,
            sheet_name="How To Use",
            index=False
        )

        # =====================================================
        # CAPITAL ALLOCATION
        # =====================================================

        print(
            "Creating Capital Allocation"
        )

        create_capital_allocation_sheet(
            writer,
            capital_allocation
        )

        # =====================================================
        # STOCK RANKINGS
        # =====================================================

        print(
            "Creating Stock Rankings"
        )

        pd.DataFrame(
            results
        ).to_excel(
            writer,
            sheet_name="Stock Rankings",
            index=False
        )

        # =====================================================
        # PORTFOLIO
        # =====================================================

        print(
            "Creating Portfolio"
        )

        portfolio_summary.to_excel(
            writer,
            sheet_name="Portfolio",
            index=False
        )

        # =====================================================
        # PORTFOLIO ACTIONS
        # =====================================================

        print(
            "Creating Portfolio Actions"
        )

        portfolio_actions.to_excel(
            writer,
            sheet_name="Portfolio Actions",
            index=False
        )

        # =====================================================
        # PORTFOLIO OPTIMISATION
        # =====================================================

        print(
            "Creating Portfolio Optimisation"
        )

        portfolio_optimisation.to_excel(
            writer,
            sheet_name="Portfolio Optimisation",
            index=False
        )

        # =====================================================
        # REBALANCE RECOMMENDATIONS
        # =====================================================

        print(
            "Creating Rebalance Recommendations"
        )

        rebalance_recommendations.to_excel(
            writer,
            sheet_name="Rebalance Recommendations",
            index=False
        )

        # =====================================================
        # SECTOR ANALYSIS
        # =====================================================

        print(
            "Creating Sector Analysis"
        )

        sector_summary.to_excel(
            writer,
            sheet_name="Sector Analysis",
            index=False
        )

        # =====================================================
        # PORTFOLIO HEALTH
        # =====================================================

        print(
            "Creating Portfolio Health"
        )

        if isinstance(
            portfolio_health,
            dict
        ):

            pd.DataFrame(
                [
                    portfolio_health
                ]
            ).to_excel(
                writer,
                sheet_name="Portfolio Health",
                index=False
            )

        elif isinstance(
            portfolio_health,
            pd.DataFrame
        ):

            portfolio_health.to_excel(
                writer,
                sheet_name="Portfolio Health",
                index=False
            )

        else:

            pd.DataFrame(
                {
                    "Status": [
                        "No portfolio health data available"
                    ]
                }
            ).to_excel(
                writer,
                sheet_name="Portfolio Health",
                index=False
            )

        # =====================================================
        # FINAL PORTFOLIO DECISIONS
        # =====================================================

        print(
            "Creating Final Portfolio Decisions"
        )

        final_portfolio_decisions.to_excel(
            writer,
            sheet_name="Final Portfolio Decisions",
            index=False
        )

        # =====================================================
        # INVESTMENT DECISIONS
        # =====================================================

        print(
            "Creating Investment Decisions"
        )

        decisions.to_excel(
            writer,
            sheet_name="Investment Decisions",
            index=False
        )

        # =====================================================
        # TRADE PLAN
        # =====================================================

        print(
            "Creating Trade Plan"
        )

        trade_plan.to_excel(
            writer,
            sheet_name="Trade Plan",
            index=False
        )

        # =====================================================
        # PORTFOLIO GROWTH PLAN
        # =====================================================

        print(
            "Creating Portfolio Growth Plan"
        )

        try:

            growth_plan_df = pd.DataFrame(
                growth_plan
            )

        except Exception:

            growth_plan_df = pd.DataFrame()

        growth_plan_df.to_excel(
            writer,
            sheet_name="Portfolio Growth Plan",
            index=False
        )

        # =====================================================
        # AI PORTFOLIO REVIEW
        # =====================================================

        print(
            "Creating AI Portfolio Review"
        )

        print(
            "AI REVIEW TYPE:",
            type(portfolio_ai_review)
        )

        if not portfolio_ai_review.empty:

            portfolio_ai_review.to_excel(
                writer,
                sheet_name="AI Portfolio Review",
                index=False
            )

        else:

            pd.DataFrame(
                {
                    "Status": [
                        "No AI portfolio review available"
                    ]
                }
            ).to_excel(
                writer,
                sheet_name="AI Portfolio Review",
                index=False
            )

        # =====================================================
        # AI PORTFOLIO MANAGER
        # =====================================================

        print(
            "Creating AI Portfolio Manager"
        )

        if isinstance(
            portfolio_manager_review,
            dict
        ):

            pd.DataFrame(
                [
                    portfolio_manager_review
                ]
            ).to_excel(
                writer,
                sheet_name="AI Portfolio Manager",
                index=False
            )

        else:

            pd.DataFrame(
                {
                    "Status": [
                        "No portfolio manager review available"
                    ]
                }
            ).to_excel(
                writer,
                sheet_name="AI Portfolio Manager",
                index=False
            )

        # =====================================================
        # RECOMMENDATION LEARNING
        # =====================================================

        print(
            "Creating Recommendation Learning"
        )

        if isinstance(
            recommendation_learning,
            dict
        ):

            # -------------------------------------------------
            # OVERALL SUMMARY
            # -------------------------------------------------

            overall = recommendation_learning.get(
                "Overall",
                {}
            )

            if isinstance(
                overall,
                dict
            ) and overall:

                pd.DataFrame(
                    [
                        overall
                    ]
                ).to_excel(
                    writer,
                    sheet_name="Recommendation Learning Summary",
                    index=False
                )

            else:

                pd.DataFrame(
                    {
                        "Status": [
                            "No recommendation learning summary available"
                        ]
                    }
                ).to_excel(
                    writer,
                    sheet_name="Recommendation Learning Summary",
                    index=False
                )

            # -------------------------------------------------
            # HORIZON LEARNING
            #
            # IMPORTANT:
            # Horizon Learning is now a DataFrame.
            # It is NOT a dictionary keyed by horizon.
            # -------------------------------------------------

            horizons = recommendation_learning.get(
                "Horizon Learning",
                pd.DataFrame()
            )

            if isinstance(
                horizons,
                pd.DataFrame
            ) and not horizons.empty:

                print(
                    "Horizon Learning rows:",
                    len(horizons)
                )

                # One consolidated sheet
                horizons.to_excel(
                    writer,
                    sheet_name="Learning Horizons",
                    index=False
                )

            else:

                pd.DataFrame(
                    {
                        "Status": [
                            "No horizon learning data available"
                        ]
                    }
                ).to_excel(
                    writer,
                    sheet_name="Learning Horizons",
                    index=False
                )

            # -------------------------------------------------
            # SIGNAL PERFORMANCE
            # -------------------------------------------------

            learning_signal = recommendation_learning.get(
                "Signal Performance",
                pd.DataFrame()
            )

            if isinstance(
                learning_signal,
                pd.DataFrame
            ) and not learning_signal.empty:

                learning_signal.to_excel(
                    writer,
                    sheet_name="Learning Signals",
                    index=False
                )

            # -------------------------------------------------
            # SIGNAL RELIABILITY
            # -------------------------------------------------

            signal_reliability = recommendation_learning.get(
                "Signal Reliability",
                pd.DataFrame()
            )

            if isinstance(
                signal_reliability,
                pd.DataFrame
            ) and not signal_reliability.empty:

                signal_reliability.to_excel(
                    writer,
                    sheet_name="Signal Reliability",
                    index=False
                )

            # -------------------------------------------------
            # SCORE BUCKET PERFORMANCE
            # -------------------------------------------------

            learning_score_bucket = (
                recommendation_learning.get(
                    "Score Bucket Performance",
                    pd.DataFrame()
                )
            )

            if isinstance(
                learning_score_bucket,
                pd.DataFrame
            ) and not learning_score_bucket.empty:

                learning_score_bucket.to_excel(
                    writer,
                    sheet_name="Learning Score Buckets",
                    index=False
                )

            # -------------------------------------------------
            # SCORE HORIZON PERFORMANCE
            # -------------------------------------------------

            score_horizon = recommendation_learning.get(
                "Score Horizon Performance",
                pd.DataFrame()
            )

            if isinstance(
                score_horizon,
                pd.DataFrame
            ) and not score_horizon.empty:

                score_horizon.to_excel(
                    writer,
                    sheet_name="Learning Score Horizons",
                    index=False
                )

            # -------------------------------------------------
            # SIGNAL HORIZON PERFORMANCE
            # -------------------------------------------------

            signal_horizon = recommendation_learning.get(
                "Signal Horizon Performance",
                pd.DataFrame()
            )

            if isinstance(
                signal_horizon,
                pd.DataFrame
            ) and not signal_horizon.empty:

                signal_horizon.to_excel(
                    writer,
                    sheet_name="Learning Signal Horizons",
                    index=False
                )

            # -------------------------------------------------
            # COMPONENT SCORE PERFORMANCE
            # -------------------------------------------------

            component_learning = (
                recommendation_learning.get(
                    "Component Score Performance",
                    pd.DataFrame()
                )
            )

            if isinstance(
                component_learning,
                pd.DataFrame
            ) and not component_learning.empty:

                component_learning.to_excel(
                    writer,
                    sheet_name="Learning Components",
                    index=False
                )

            # -------------------------------------------------
            # CONFIDENCE PERFORMANCE
            # -------------------------------------------------

            confidence_learning = (
                recommendation_learning.get(
                    "Confidence Performance",
                    pd.DataFrame()
                )
            )

            if isinstance(
                confidence_learning,
                pd.DataFrame
            ) and not confidence_learning.empty:

                confidence_learning.to_excel(
                    writer,
                    sheet_name="Learning Confidence",
                    index=False
                )

        else:

            pd.DataFrame(
                {
                    "Status": [
                        "No recommendation learning available"
                    ]
                }
            ).to_excel(
                writer,
                sheet_name="Recommendation Learning Summary",
                index=False
            )

        # =====================================================
        # RECOMMENDATION INTELLIGENCE
        # =====================================================

        print(
            "Creating Recommendation Intelligence"
        )

        intelligence_sheet = (
            "Recommendation Intelligence"
        )

        # Always create the sheet
        pd.DataFrame(
            {
                "Status": [
                    "Recommendation Intelligence Report"
                ]
            }
        ).to_excel(
            writer,
            sheet_name=intelligence_sheet,
            index=False
        )

        intelligence_row = 2

        intelligence_sections = [

            (
                "Signal Performance",
                signal_performance
            ),

            (
                "Horizon Performance",
                horizon_performance
            ),

            (
                "Recommendation Intelligence",
                recommendation_intelligence
            ),

            (
                "Score Performance",
                score_performance
            ),

            (
                "Score Bucket Performance",
                score_bucket_performance
            ),

            (
                "Component Score Performance",
                component_score_performance
            ),

            (
                "Signal Horizon Performance",
                signal_horizon_performance
            )

        ]

        for title, dataframe in intelligence_sections:

            if dataframe is None:
                continue

            if not isinstance(
                dataframe,
                pd.DataFrame
            ):
                continue

            if dataframe.empty:
                continue

            pd.DataFrame(
                [
                    [
                        title
                    ]
                ]
            ).to_excel(
                writer,
                sheet_name=intelligence_sheet,
                startrow=intelligence_row,
                index=False,
                header=False
            )

            intelligence_row += 1

            dataframe.to_excel(
                writer,
                sheet_name=intelligence_sheet,
                startrow=intelligence_row,
                index=False
            )

            intelligence_row += (
                len(dataframe) + 4
            )

        # =====================================================
        # ALERTS
        # =====================================================

        print(
            "ALERTS SHAPE:",
            alerts.shape
        )

        print(
            "ALERTS EMPTY:",
            alerts.empty
        )

        print(
            "Creating Alerts"
        )

        alerts.to_excel(
            writer,
            sheet_name="Alerts",
            index=False
        )

    # =========================================================
    # WORKBOOK VERIFICATION
    # =========================================================

    print(
        "Report created:",
        filename
    )

    try:

        from openpyxl import load_workbook

        workbook = load_workbook(
            filename,
            read_only=True
        )

        print(
            "FINAL WORKBOOK SHEETS:",
            workbook.sheetnames
        )

        workbook.close()

    except Exception as e:

        print(
            "Workbook verification failed:",
            e
        )

    return filename
